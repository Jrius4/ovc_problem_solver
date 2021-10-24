from openpyxl import load_workbook

wb = load_workbook("input/input.xlsx")
ws = wb.active
data = []

def getcolumns(a:str,b:str):
    a = str(a).lower()
    b = str(b).lower()
    if a == "":
        return ((26*(0))+int(ord(b)-97))
    else:
        return ((26*(int(ord(a)-97)+1))+int(ord(b)-97))
    



for row in ws.iter_rows(min_col=1,min_row= 3,max_col=ws.max_column,max_row=ws.max_row,values_only=True):
    data.append(row)
print("\n\n")
print(data[0])
print("\n\n")
print(data[0][getcolumns("b","e")])
print("\n\n")

print(len(data))
print("\n\n")

# no of ovc (0-17) with HIV Positive Caregiver
hh_with_positve_cg = []
hh_no_with_positve_cg = []
for i in data:
    if i[getcolumns("b","e")] == "HIV Positive" and i[getcolumns("","l")] == i[getcolumns("","o")]:
        hh_with_positve_cg.append([i[getcolumns("","m")],i[getcolumns("","l")],i[getcolumns("","o")],i[getcolumns("c","m")]])
        hh_no_with_positve_cg.append(i[getcolumns("c","m")])
print("\n\n")
print(hh_with_positve_cg)
print("\n\n")
print(hh_no_with_positve_cg)

y_ages = []
for age in hh_with_positve_cg:
    if age[0] <=17:
     y_ages.append(age)   

print("\n\n")
print(f"\n\n hh of 17 yrs")

print(y_ages)
print("\n\n")


# filter out ovcs from -->> no of ovc (0-17) with HIV Positive Caregiver
# 571 ovc with positive cg
ovc_with_positive_cg = []
for i in data:
    if i[getcolumns("","m")] <= 17 and i[getcolumns("c","m")] in hh_no_with_positve_cg:
        ovc_with_positive_cg.append([i[getcolumns("","m")],i[getcolumns("","l")],i[getcolumns("","o")],i[getcolumns("c","m")]])
print("\n\n ovc_with_positive_cg ")       
print(ovc_with_positive_cg)
print("\n\n")
print("\n\n")
print(len(ovc_with_positive_cg))
k = []
for i in ovc_with_positive_cg:
    if i[2] == 'NALWEYISO DIANA':
        k.append(i)

print("\n\n 'NALWEYISO DIANA' ")       
print(k)
print("\n\n")
print("\n\n")
print(len(k))

print("\n\n")
print("\n\n")

###########################################################
############################################################
############################################################



# no hh with hiv positve cg and clhiv (<18)
hh_with_positve_pp_lhiv = []
hh_no_with_positve_pp_lhiv = []
for i in data:
    if i[getcolumns("b","e")] == "HIV Positive":
        hh_with_positve_pp_lhiv.append([i[getcolumns("","m")],i[getcolumns("","l")],i[getcolumns("","o")],i[getcolumns("c","m")]])
        hh_no_with_positve_pp_lhiv.append(i[getcolumns("c","m")])


print("\n\n hh with  hh_with_positve_pp_lhiv")       
print(hh_with_positve_pp_lhiv)
print("\n\n")
print("\n\n")
print(len(hh_with_positve_pp_lhiv))
print("\n\n")
print(hh_no_with_positve_pp_lhiv)
print("\n\n")
print(len(hh_no_with_positve_pp_lhiv))
print(sorted(hh_no_with_positve_pp_lhiv))

result = []

for i in hh_no_with_positve_pp_lhiv:
    if i not in result:
        result.append(i)

print("\n\n")
print(sorted(result))
result_2 = result
for i in result_2:
    if i not in hh_no_with_positve_cg:
        result_2.remove(i)
print("\n\n")
print(sorted(result_2))
print(len(sorted(result_2)))

set_all_pos = set(result)
set_all_pos_cg = set(hh_no_with_positve_cg)

set_intersec = set_all_pos.intersection(set_all_pos_cg)
print(len(set_intersec))